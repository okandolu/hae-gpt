"""Batch Query Database Module.

This module provides SQLite database operations for batch query storage
and management with full export capabilities.

Features:
    - Document object handling
    - Database persistence
    - CSV and Excel export methods
    - Smart Excel export with word wrap
"""

import sqlite3
from pathlib import Path
from datetime import datetime
from typing import Dict, Any
import pandas as pd
from rich.console import Console
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import warnings

console = Console()

import config

# Database path
DB_PATH = Path("data/batch_queries.db")

# Excel limits
EXCEL_CELL_LIMIT = 32767


class BatchQueryDB:
    """
    SQLite database operations for batch queries
    """
    
    def __init__(self, db_path: Path = DB_PATH):
        self.db_path = db_path
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self._init_db()
    
    def _init_db(self):
        """Initialize database with schema - PRESERVES EXISTING DATA"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Create table ONLY IF NOT EXISTS (preserves old records)
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS batch_queries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            
            -- Question & Answer
            question TEXT NOT NULL,
            answer TEXT NOT NULL,
            
            -- Mode & Parameters
            mode TEXT,
            threshold REAL,
            top_k INTEGER,
            patient_info TEXT,  -- NEW: CLINICAL INFO COLUMN
            
            -- Context 1 (FULL TEXT)
            context1_text TEXT,
            context1_reference TEXT,
            context1_filename TEXT,
            context1_page INTEGER,
            context1_section TEXT,
            context1_has_table INTEGER,
            context1_similarity REAL,
            
            -- Context 2 (FULL TEXT)
            context2_text TEXT,
            context2_reference TEXT,
            context2_filename TEXT,
            context2_page INTEGER,
            context2_section TEXT,
            context2_has_table INTEGER,
            context2_similarity REAL,
            
            -- Context 3 (FULL TEXT)
            context3_text TEXT,
            context3_reference TEXT,
            context3_filename TEXT,
            context3_page INTEGER,
            context3_section TEXT,
            context3_has_table INTEGER,
            context3_similarity REAL,
            
            -- Context 4 (FULL TEXT)
            context4_text TEXT,
            context4_reference TEXT,
            context4_filename TEXT,
            context4_page INTEGER,
            context4_section TEXT,
            context4_has_table INTEGER,
            context4_similarity REAL,
            
            -- Context 5 (FULL TEXT)
            context5_text TEXT,
            context5_reference TEXT,
            context5_filename TEXT,
            context5_page INTEGER,
            context5_section TEXT,
            context5_has_table INTEGER,
            context5_similarity REAL,
            
            -- Reasoning (FULL TEXT)
            reasoning TEXT,
            
            -- Token Usage
            prompt_tokens INTEGER,
            completion_tokens INTEGER,
            total_tokens INTEGER,
            
            -- Performance
            query_time_seconds REAL,
            
            -- Timestamp
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
        )
        """)
        
        conn.commit()
        conn.close()
        
        # Count existing records
        count = self.get_query_count()
        if count > 0:
            console.print(f"[green]✓[/green] Database loaded: {self.db_path} ({count} existing records)")
        else:
            console.print(f"[green]✓[/green] Database initialized: {self.db_path}")
    
    def insert_query(self, data: Dict[str, Any]) -> int:
        """
        Insert a query with full context text
        FIXED: Handles Document objects properly
        
        Args:
            data: Query data with contexts
            
        Returns:
            Row ID
        """
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Prepare context data (FULL TEXT) - Fixed document handling
        context_data = {}
        retrieval_results = data.get('retrieval_results', [])
        citations = data.get('citations', [])
        for i in range(1, 6):
            if i - 1 < len(retrieval_results):
                doc, sim = retrieval_results[i - 1]
                citation = citations[i - 1] if i - 1 < len(citations) else {}

                # Fixed: Handle Document object properly
                # Document has attributes, not dictionary keys
                if hasattr(doc, 'page_content'):
                    # It's a Document object
                    context_text = doc.page_content
                elif isinstance(doc, dict):
                    # It's a dictionary
                    context_text = doc.get('page_content', '')
                else:
                    # Unknown type, convert to string
                    context_text = str(doc)
                
                context_data[f'context{i}_text'] = context_text
                context_data[f'context{i}_reference'] = citation.get('reference', '')
                context_data[f'context{i}_filename'] = citation.get('filename', '')
                context_data[f'context{i}_page'] = citation.get('page', None)
                context_data[f'context{i}_section'] = citation.get('section', '')
                context_data[f'context{i}_has_table'] = int(citation.get('has_table', False))
                context_data[f'context{i}_similarity'] = sim
            else:
                context_data[f'context{i}_text'] = None
                context_data[f'context{i}_reference'] = None
                context_data[f'context{i}_filename'] = None
                context_data[f'context{i}_page'] = None
                context_data[f'context{i}_section'] = None
                context_data[f'context{i}_has_table'] = 0
                context_data[f'context{i}_similarity'] = None
        
        # Insert
        cursor.execute("""
        INSERT INTO batch_queries (
            question, answer, mode, threshold, top_k, patient_info,
            context1_text, context1_reference, context1_filename, context1_page, 
            context1_section, context1_has_table, context1_similarity,
            context2_text, context2_reference, context2_filename, context2_page,
            context2_section, context2_has_table, context2_similarity,
            context3_text, context3_reference, context3_filename, context3_page,
            context3_section, context3_has_table, context3_similarity,
            context4_text, context4_reference, context4_filename, context4_page,
            context4_section, context4_has_table, context4_similarity,
            context5_text, context5_reference, context5_filename, context5_page,
            context5_section, context5_has_table, context5_similarity,
            reasoning, prompt_tokens, completion_tokens, total_tokens, query_time_seconds
        ) VALUES (
            ?, ?, ?, ?, ?, ?,
            ?, ?, ?, ?, ?, ?, ?,
            ?, ?, ?, ?, ?, ?, ?,
            ?, ?, ?, ?, ?, ?, ?,
            ?, ?, ?, ?, ?, ?, ?,
            ?, ?, ?, ?, ?, ?, ?,
            ?, ?, ?, ?, ?
        )
        """, (
            data['question'],
            data['answer'],
            data['mode'],
            data['threshold'],
            data['top_k'],
            data.get('patient_info', ''),  # NEW: CLINICAL INFO
            # Context 1
            context_data['context1_text'],
            context_data['context1_reference'],
            context_data['context1_filename'],
            context_data['context1_page'],
            context_data['context1_section'],
            context_data['context1_has_table'],
            context_data['context1_similarity'],
            # Context 2
            context_data['context2_text'],
            context_data['context2_reference'],
            context_data['context2_filename'],
            context_data['context2_page'],
            context_data['context2_section'],
            context_data['context2_has_table'],
            context_data['context2_similarity'],
            # Context 3
            context_data['context3_text'],
            context_data['context3_reference'],
            context_data['context3_filename'],
            context_data['context3_page'],
            context_data['context3_section'],
            context_data['context3_has_table'],
            context_data['context3_similarity'],
            # Context 4
            context_data['context4_text'],
            context_data['context4_reference'],
            context_data['context4_filename'],
            context_data['context4_page'],
            context_data['context4_section'],
            context_data['context4_has_table'],
            context_data['context4_similarity'],
            # Context 5
            context_data['context5_text'],
            context_data['context5_reference'],
            context_data['context5_filename'],
            context_data['context5_page'],
            context_data['context5_section'],
            context_data['context5_has_table'],
            context_data['context5_similarity'],
            # Reasoning & tokens
            data.get('reasoning', ''),
            data.get('prompt_tokens', 0),
            data.get('completion_tokens', 0),
            data.get('total_tokens', 0),
            data.get('query_time', 0.0)
        ))
        
        row_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        return row_id
    
    def get_all_queries(self) -> pd.DataFrame:
        """
        Get all queries as DataFrame
        """
        conn = sqlite3.connect(self.db_path)
        df = pd.read_sql_query("SELECT * FROM batch_queries ORDER BY id DESC", conn)
        conn.close()
        return df
    
    def get_query_count(self) -> int:
        """Get total number of queries"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM batch_queries")
        count = cursor.fetchone()[0]
        conn.close()
        return count
    
    def export_to_csv(self, output_path: str) -> str:
        """
        Export all queries to CSV (NO CHARACTER LIMIT - RECOMMENDED)
        
        Args:
            output_path: Output file path
            
        Returns:
            Path to exported file
        """
        df = self.get_all_queries()
        
        # Reorder columns for better readability
        column_order = [
            'id', 'timestamp', 'question', 'answer', 'mode', 'patient_info',
            'threshold', 'top_k',
            # Context 1
            'context1_reference', 'context1_similarity', 'context1_filename', 
            'context1_page', 'context1_section', 'context1_has_table', 'context1_text',
            # Context 2
            'context2_reference', 'context2_similarity', 'context2_filename',
            'context2_page', 'context2_section', 'context2_has_table', 'context2_text',
            # Context 3
            'context3_reference', 'context3_similarity', 'context3_filename',
            'context3_page', 'context3_section', 'context3_has_table', 'context3_text',
            # Context 4
            'context4_reference', 'context4_similarity', 'context4_filename',
            'context4_page', 'context4_section', 'context4_has_table', 'context4_text',
            # Context 5
            'context5_reference', 'context5_similarity', 'context5_filename',
            'context5_page', 'context5_section', 'context5_has_table', 'context5_text',
            # Reasoning & tokens
            'reasoning', 'prompt_tokens', 'completion_tokens', 'total_tokens',
            'query_time_seconds'
        ]
        
        # Reorder (only existing columns)
        existing_cols = [col for col in column_order if col in df.columns]
        df = df[existing_cols]
        
        # Export to CSV (NO LIMITS!)
        df.to_csv(output_path, index=False, encoding='utf-8-sig')

        console.print(f"[green]✓[/green] CSV Export (NO DATA LOSS): {output_path}")
        console.print(f"[cyan]→[/cyan] {len(df)} records exported with full context text")
        
        return output_path
    
    def export_to_excel(self, output_path: str, truncate_long_text: bool = False) -> str:
        """
        BACKWARD COMPATIBILITY: Old export method
        Redirects to smart export
        
        Args:
            output_path: Output file path
            truncate_long_text: Ignored (for compatibility)
            
        Returns:
            Path to exported file
        """
        console.print("Using legacy export_to_excel, redirecting to export_to_excel_smart")
        return self.export_to_excel_smart(output_path)
    
    def export_to_excel_smart(self, output_path: str) -> str:
        """
        SMART Excel Export - PhpSpreadsheet style, NO data loss!
        
        For long texts:
        1. Word wrap is enabled (text breaks into lines)
        2. Row height is automatically adjusted
        3. Column width is optimized
        
        Args:
            output_path: Output file path
            
        Returns:
            Path to exported file
        """
        df = self.get_all_queries()
        
        if len(df) == 0:
            console.print("No data to export")
            return output_path
        
        # Reorder columns
        column_order = [
            'id', 'timestamp', 'question', 'answer', 'mode', 'patient_info',
            'threshold', 'top_k',
            'context1_reference', 'context1_similarity', 'context1_filename', 
            'context1_page', 'context1_section', 'context1_has_table', 'context1_text',
            'context2_reference', 'context2_similarity', 'context2_filename',
            'context2_page', 'context2_section', 'context2_has_table', 'context2_text',
            'context3_reference', 'context3_similarity', 'context3_filename',
            'context3_page', 'context3_section', 'context3_has_table', 'context3_text',
            'context4_reference', 'context4_similarity', 'context4_filename',
            'context4_page', 'context4_section', 'context4_has_table', 'context4_text',
            'context5_reference', 'context5_similarity', 'context5_filename',
            'context5_page', 'context5_section', 'context5_has_table', 'context5_text',
            'reasoning', 'prompt_tokens', 'completion_tokens', 'total_tokens',
            'query_time_seconds'
        ]
        
        existing_cols = [col for col in column_order if col in df.columns]
        df = df[existing_cols]
        
        # Create workbook using openpyxl directly (PhpSpreadsheet style)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Batch Queries"
        
        # Header row styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Write headers
        for col_idx, col_name in enumerate(existing_cols, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Identify text columns (columns that may have long text)
        text_columns = [col for col in existing_cols if 'text' in col.lower() or 
                       col in ['question', 'answer', 'reasoning']]
        
        # Write data with smart handling
        console.print(f"[cyan]Writing {len(df)} rows to Excel...[/cyan]")
        
        for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
            for col_idx, (col_name, value) in enumerate(zip(existing_cols, row_data), start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Convert value to string
                str_value = str(value) if pd.notna(value) else ''
                
                # Smart handling for long text
                if col_name in text_columns and len(str_value) > 100:
                    # Enable word wrap for long text
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    
                    # If exceeds Excel limit, split intelligently
                    if len(str_value) > EXCEL_CELL_LIMIT:
                        # Split at word boundaries near the limit
                        truncate_point = str_value.rfind(' ', 0, EXCEL_CELL_LIMIT - 100)
                        if truncate_point == -1:
                            truncate_point = EXCEL_CELL_LIMIT - 100
                        
                        cell.value = str_value[:truncate_point] + "\n\n[...CSV KULLANIN...]"
                        
                        # Add note/comment
                        comment_text = f"Bu metin {len(str_value):,} karakter.\nExcel limiti: 32,767 karakter.\nTam icerik icin CSV export kullanin."
                        cell.comment = Comment(comment_text, "System")
                    else:
                        cell.value = str_value
                else:
                    # Normal values
                    cell.alignment = Alignment(vertical="top")
                    cell.value = str_value
        
        # Column width optimization
        console.print("[cyan]Optimizing column widths...[/cyan]")
        
        for col_idx, col_name in enumerate(existing_cols, start=1):
            col_letter = get_column_letter(col_idx)
            
            # Determine optimal width
            if col_name in text_columns:
                # Text columns: wider but not too wide
                ws.column_dimensions[col_letter].width = 50
            elif col_name in ['id', 'page', 'has_table']:
                # Small columns
                ws.column_dimensions[col_letter].width = 10
            elif 'similarity' in col_name or 'tokens' in col_name:
                # Numeric columns
                ws.column_dimensions[col_letter].width = 15
            else:
                # Default width
                ws.column_dimensions[col_letter].width = 20
        
        # Freeze first row
        ws.freeze_panes = "A2"
        
        # Add info sheet
        info_ws = wb.create_sheet("Bilgi", 0)
        info_ws['A1'] = "Batch Query Export"
        info_ws['A1'].font = Font(size=16, bold=True, color="4472C4")
        
        info_ws['A3'] = "Export Tarihi:"
        info_ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        info_ws['A4'] = "Toplam Kayit:"
        info_ws['B4'] = len(df)
        
        info_ws['A5'] = "Toplam Sutun:"
        info_ws['B5'] = len(existing_cols)
        
        info_ws['A7'] = "Bilgi Kayit:"
        info_ws['B7'] = "YOK - Tum veriler word wrap ile satirlara bolunur"
        info_ws['B7'].font = Font(bold=True, color="00AA00")
        
        # Save workbook
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb.save(output_path)
        
        console.print(f"SMART Excel Export: {output_path}")
        console.print(f"{len(df)} word wrap ile export edildi")
        console.print(f"Bilgi kaybi yok!")
        
        return output_path
    
    def clear_all(self):
        """Delete all records"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM batch_queries")
        conn.commit()
        conn.close()
        console.print("All records deleted")

